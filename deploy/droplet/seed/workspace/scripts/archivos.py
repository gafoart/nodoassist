#!/usr/bin/env python3
"""Archivos del cliente: guardar entrantes con etiqueta natural y recuperarlos.

Los archivos entrantes persisten en <state>/media/inbound/ con nombre
"<orig>---<uuid>.ext" pero no hay forma de recuperarlos por referencia natural
("la lista de precios de maria"). Este script mantiene un indice propio en
<workspace>/archivos/ (copia + indice.json + INDICE.md).

Uso:
  archivos.py guardar "<etiqueta>" (--ultimo-inbound | --ruta PATH) [--de QUIEN]
  archivos.py buscar "<texto>"
  archivos.py leer "<etiqueta u ordinal>" [--todo] [--hoja N]
  archivos.py listar

Lee xlsx (openpyxl), xls (xlrd), csv (stdlib) y PDF (pdftotext -layout).
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import shutil
import subprocess
import sys
import unicodedata
from datetime import datetime
from pathlib import Path

STATE_DIR = Path(
    __import__("os").environ.get("NODOASSIST_STATE_DIR", str(Path.home() / ".nodoassist"))
)
WORKSPACE = Path(
    __import__("os").environ.get("NODOASSIST_WORKSPACE_DIR", str(STATE_DIR / "workspace"))
)
INBOUND = STATE_DIR / "media" / "inbound"
ARCHIVE = WORKSPACE / "archivos"
INDEX_JSON = ARCHIVE / "indice.json"
INDEX_MD = ARCHIVE / "INDICE.md"

MAX_PREVIEW_ROWS = 40
MAX_CELL_CHARS = 80


def normalize(text: str) -> str:
    text = unicodedata.normalize("NFD", text.lower())
    return "".join(ch for ch in text if unicodedata.category(ch) != "Mn")


def slugify(label: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", normalize(label)).strip("-")
    return slug or "archivo"


def load_index() -> list[dict]:
    try:
        return json.loads(INDEX_JSON.read_text(encoding="utf-8"))
    except (OSError, ValueError):
        return []


def write_index(entries: list[dict]) -> None:
    ARCHIVE.mkdir(parents=True, exist_ok=True)
    INDEX_JSON.write_text(json.dumps(entries, ensure_ascii=False, indent=2), encoding="utf-8")
    lines = ["# Índice de archivos del cliente", ""]
    for i, e in enumerate(entries, 1):
        de = f" · de {e['de']}" if e.get("de") else ""
        lines.append(f"{i}. **{e['etiqueta']}**{de} · {e['fecha']} · `{e['archivo']}` (orig: {e['original']})")
    INDEX_MD.write_text("\n".join(lines) + "\n", encoding="utf-8")


def latest_inbound() -> Path:
    files = [p for p in INBOUND.glob("*") if p.is_file()]
    if not files:
        sys.exit(f"No hay archivos en {INBOUND}")
    return max(files, key=lambda p: p.stat().st_mtime)


def original_name(path: Path) -> str:
    # inbound files are stored as "<orig>---<uuid>.ext"
    m = re.match(r"^(.*)---[0-9a-fA-F-]{8,}(\.[^.]+)?$", path.name)
    return (m.group(1) + (m.group(2) or "")) if m else path.name


def cmd_guardar(args: argparse.Namespace) -> None:
    src = Path(args.ruta) if args.ruta else latest_inbound()
    if not src.is_file():
        sys.exit(f"No existe: {src}")
    entries = load_index()
    slug = slugify(args.etiqueta)
    dest = ARCHIVE / f"{slug}{src.suffix.lower()}"
    n = 2
    while dest.exists():
        dest = ARCHIVE / f"{slug}-{n}{src.suffix.lower()}"
        n += 1
    ARCHIVE.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dest)
    entries.append(
        {
            "etiqueta": args.etiqueta,
            "archivo": dest.name,
            "original": original_name(src),
            "de": args.de or "",
            "fecha": datetime.now().strftime("%Y-%m-%d %H:%M"),
        }
    )
    write_index(entries)
    print(f"Guardado: «{args.etiqueta}» → archivos/{dest.name} (orig: {original_name(src)})")


def match_entries(query: str, entries: list[dict]) -> list[tuple[int, dict]]:
    tokens = [t for t in normalize(query).split() if t]
    scored = []
    for i, e in enumerate(entries):
        haystack = normalize(" ".join([e["etiqueta"], e["original"], e.get("de", "")]))
        hits = sum(1 for t in tokens if t in haystack)
        if hits:
            scored.append((hits, i, e))
    scored.sort(key=lambda x: (-x[0], -x[1]))
    return [(i, e) for _, i, e in scored]


def cmd_buscar(args: argparse.Namespace) -> None:
    entries = load_index()
    results = match_entries(args.texto, entries)
    if not results:
        print("Sin resultados. Usa `listar` para ver todo.")
        return
    for i, e in results[:10]:
        de = f" · de {e['de']}" if e.get("de") else ""
        print(f"{i + 1}. {e['etiqueta']}{de} · {e['fecha']} · archivos/{e['archivo']}")


def cmd_listar(_args: argparse.Namespace) -> None:
    entries = load_index()
    if not entries:
        print("Índice vacío.")
        return
    for i, e in enumerate(entries, 1):
        de = f" · de {e['de']}" if e.get("de") else ""
        print(f"{i}. {e['etiqueta']}{de} · {e['fecha']} · archivos/{e['archivo']}")


def clip(value: object) -> str:
    text = "" if value is None else str(value)
    return text if len(text) <= MAX_CELL_CHARS else text[: MAX_CELL_CHARS - 1] + "…"


def read_xlsx(path: Path, sheet: int | None, todo: bool) -> None:
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    names = wb.sheetnames
    picked = names[sheet - 1] if sheet else names[0]
    print(f"Hojas: {', '.join(names)} — mostrando «{picked}»")
    ws = wb[picked]
    limit = None if todo else MAX_PREVIEW_ROWS
    for row_index, row in enumerate(ws.iter_rows(values_only=True), 1):
        if limit is not None and row_index > limit:
            print(f"… (usa --todo para las {ws.max_row} filas)")
            break
        print(" | ".join(clip(c) for c in row))


def read_xls(path: Path, sheet: int | None, todo: bool) -> None:
    import xlrd

    wb = xlrd.open_workbook(str(path))
    names = wb.sheet_names()
    picked = names[sheet - 1] if sheet else names[0]
    print(f"Hojas: {', '.join(names)} — mostrando «{picked}»")
    ws = wb.sheet_by_name(picked)
    limit = ws.nrows if todo else min(ws.nrows, MAX_PREVIEW_ROWS)
    for r in range(limit):
        print(" | ".join(clip(c.value) for c in ws.row(r)))
    if limit < ws.nrows:
        print(f"… (usa --todo para las {ws.nrows} filas)")


def read_csv(path: Path, todo: bool) -> None:
    with path.open(newline="", encoding="utf-8", errors="replace") as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=",;\t")
        except csv.Error:
            dialect = csv.excel
        reader = csv.reader(f, dialect)
        for i, row in enumerate(reader, 1):
            if not todo and i > MAX_PREVIEW_ROWS:
                print("… (usa --todo para el resto)")
                break
            print(" | ".join(clip(c) for c in row))


def read_pdf(path: Path, todo: bool) -> None:
    result = subprocess.run(
        ["pdftotext", "-layout", str(path), "-"], capture_output=True, text=True
    )
    if result.returncode != 0:
        sys.exit(f"pdftotext falló: {result.stderr.strip()}")
    lines = result.stdout.splitlines()
    limit = len(lines) if todo else min(len(lines), MAX_PREVIEW_ROWS * 2)
    print("\n".join(lines[:limit]))
    if limit < len(lines):
        print(f"… (usa --todo para las {len(lines)} líneas)")


def cmd_leer(args: argparse.Namespace) -> None:
    entries = load_index()
    entry = None
    if args.etiqueta.isdigit() and 1 <= int(args.etiqueta) <= len(entries):
        entry = entries[int(args.etiqueta) - 1]
    else:
        results = match_entries(args.etiqueta, entries)
        entry = results[0][1] if results else None
    if not entry:
        sys.exit("No encontrado. Usa `listar` o `buscar`.")
    path = ARCHIVE / entry["archivo"]
    if not path.is_file():
        sys.exit(f"El archivo del índice no existe: {path}")
    print(f"— {entry['etiqueta']} ({entry['original']}) —")
    suffix = path.suffix.lower()
    if suffix == ".xlsx":
        read_xlsx(path, args.hoja, args.todo)
    elif suffix == ".xls":
        read_xls(path, args.hoja, args.todo)
    elif suffix in (".csv", ".tsv"):
        read_csv(path, args.todo)
    elif suffix == ".pdf":
        read_pdf(path, args.todo)
    else:
        text = path.read_text(encoding="utf-8", errors="replace")
        print(text if args.todo else text[:8000])


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    sub = parser.add_subparsers(dest="cmd", required=True)

    g = sub.add_parser("guardar")
    g.add_argument("etiqueta")
    g.add_argument("--ultimo-inbound", action="store_true")
    g.add_argument("--ruta")
    g.add_argument("--de", default="")
    g.set_defaults(fn=cmd_guardar)

    b = sub.add_parser("buscar")
    b.add_argument("texto")
    b.set_defaults(fn=cmd_buscar)

    r = sub.add_parser("leer")
    r.add_argument("etiqueta")
    r.add_argument("--todo", action="store_true")
    r.add_argument("--hoja", type=int)
    r.set_defaults(fn=cmd_leer)

    ls = sub.add_parser("listar")
    ls.set_defaults(fn=cmd_listar)

    args = parser.parse_args()
    args.fn(args)


if __name__ == "__main__":
    main()
